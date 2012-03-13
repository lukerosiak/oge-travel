from openpyxl import load_workbook
from openpyxl.shared.exc import NamedRangeException, InvalidFileException
import xlrd
import urllib2
import csv
import os

"""
Download all the excel spreadsheets representing various incarnations of a travel disclosure form from ethics.gov, and concatenate them into one big CSV.
Note that there are PDFs, too, which are digital and could have pdftotext run on them, but you'd have to parse out the layout into a CSV-like object
first. This utility omits the PDFs.
"""

def main():
    os.system("mkdir files")
    
    raw = urllib2.urlopen("http://explore.data.gov/api/views/kxfh-um2n/rows.csv")
    index = csv.DictReader(raw)

    out = csv.writer(open('all.csv','w'))

    for row in index:
        id = row['Report URL'].split('=')[1]
        
        meta = [row['Report URL'],row['Agency'],row['Filing Date'],row['Period Start'],row['Period End']]
        
        if row['Report File Type'].lower()=='xlsx':

            if not os.path.exists('files/%s.xlsx' % id):
                os.system('wget %s -O files/%s.xlsx' % (row['Report URL'],id))
            
            try: #a couple documents get a weird openpyxl error when opening
                wb = load_workbook(filename = 'files/%s.xlsx' % id, use_iterators = True)
                for ws in wb.worksheets:
                    if ws.title in ['Instruction Sheet','Agency Acronym']:
                        continue
                    for rawrow in ws.iter_rows(): 
                        row = []
                        for cell in rawrow:
                            try:
                                row.append( cell.internal_value.encode('ascii','ignore') )
                            except:
                                row.append( cell.internal_value )
                        out.writerow([id+'.xlsx',ws.title] + meta + row)
            except NamedRangeException:
                print "couldnt open %s" % id
                pass
            except InvalidFileException:
                print "couldnt open %s" % id
                pass               
                
        elif row['Report File Type'].lower()=='xls':
            
            if not os.path.exists('files/%s.xls' % id):
                os.system('wget %s -O files/%s.xls' % (row['Report URL'],id))
        
            book = xlrd.open_workbook('files/%s.xls' % id) 
            for ws in book.sheets():
                for i in xrange(ws.nrows):
                    rawrow = ws.row(i)
                    row = []
                    for cell in rawrow:
                        if cell.ctype==xlrd.XL_CELL_DATE:
                            xlrd_date = xlrd.xldate_as_tuple(cell.value, book.datemode)
                            row.append( "%s-%s-%s" % (xlrd_date[0], xlrd_date[1], xlrd_date[2]) )
                        elif cell.ctype==xlrd.XL_CELL_TEXT:
                            row.append( cell.value.encode('ascii','ignore') )
                        else:
                            row.append(cell.value)
                    out.writerow([id+'.xls',ws.name] + meta + row)         

main()
